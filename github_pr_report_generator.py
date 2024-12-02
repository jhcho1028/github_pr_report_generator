import requests
import pandas as pd
import time
import json
import os
from os.path import exists
from requests.exceptions import ConnectionError, Timeout, RequestException
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import Alignment, Font
from datetime import datetime

# GitHub Personal Access Token
GITHUB_TOKEN = os.environ.get('JH_TOKEN')
if not GITHUB_TOKEN:
    raise EnvironmentError("GITHUB_TOKEN environment variable not set.")

# GitHub API request headers
headers = {
    'Authorization': f'token {GITHUB_TOKEN}',
    'Accept': 'application/vnd.github.v3+json'
}

CACHE_FILE = "pr_cache.json"

# Load cache
def load_cache():
    if exists(CACHE_FILE):
        with open(CACHE_FILE, "r") as file:
            return json.load(file)
    return {}

# Save cache
def save_cache(cache):
    with open(CACHE_FILE, "w") as file:
        json.dump(cache, file)

pr_cache = load_cache()

def get_rate_limit():
    """Fetch GitHub API rate limit."""
    rate_limit_url = 'https://api.github.com/rate_limit'
    rate_limit_response = requests.get(rate_limit_url, headers=headers)
    print(json.dumps(rate_limit_response.json(), indent=2))

def ensure_rate_limit():
    rate_limit_url = 'https://api.github.com/rate_limit'
    response = requests.get(rate_limit_url, headers=headers)
    rate_limit_data = response.json()

    remaining = rate_limit_data['rate']['remaining']    # 남은 요청 수
    reset_time = rate_limit_data['rate']['reset']   # 리셋 시간 

    if remaining == 0:
        # 요청이 초과되었을 경우, 리셋 될 때 까지 대기
        wait_time = reset_time - int(time.time())
        print(f"Rate limit exceeded. Waiting for {wait_time} seconds.")
        time.sleep(wait_time + 1)

def get_user_id(username):
    """Fetch the user ID from GitHub API based on username."""
    ensure_rate_limit()
    url = f'https://api.github.com/users/{username}'
    response = requests.get(url, headers=headers)
    
    if response.status_code == 200:
        user_info = response.json()
        return user_info['id']  # Get the unique user ID
    else:
        print(f"Error fetching user info for {username}")
        return None
    
def get_repositories_from_excel(repo_excel_path, repo_sheet_name, column_letter):
    """Load repositories from the Excel file's specified column."""
    try:
        wb = load_workbook(repo_excel_path)
        ws = wb[repo_sheet_name]

        repos = []
        column_index = column_index_from_string(column_letter) - 1

        for row in ws.iter_rows(min_row=2, values_only=True):  # No header # Skip header row
            repo_name = row[column_index]
            print(f"Repo name: '{repo_name}'")
            if repo_name and repo_name.strip():
                repos.append(repo_name.strip())  # 이름 정리
            else:
                print("Skipping empty repository name")

        print(f"Found {len(repos)} repositories.")
        return repos
    except FileNotFoundError:
        print(f"Error: File not found at {repo_excel_path}")
        return []
    except KeyError:
        print(f"Error: Sheet '{repo_sheet_name}' not found in the Excel file.")
        return []
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        return []

def get_contributors_from_excel(repo_excel_path, contributor_sheet_name, column_letter):
    """Load contributors from the Excel file's specified column."""
    print(f"Loading contributors from Excel file '{repo_excel_path}', sheet '{contributor_sheet_name}'...")
    wb = load_workbook(repo_excel_path)
    ws = wb[contributor_sheet_name]

    contributors = []
    column_index = column_index_from_string(column_letter) - 1

    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header row
        contributor_name = row[column_index]
        if contributor_name:
            contributors.append(contributor_name)
        else:
            print("Skipping empty contributor name")

    print(f"Found {len(contributors)} contributors.")
    return contributors

def get_prs_for_repository(repo_name):
    """Fetch PRs for a given repository."""
    print(f"Fetching PR list for repository '{repo_name}'...")
    all_prs = []
    page = 1

    while True:
        url = f'https://api.github.com/repos/AdvancedTechnologyInc/{repo_name}/pulls?state=all&per_page=100&page={page}'
        try:
            ensure_rate_limit()
            response = requests.get(url, headers=headers, timeout=10)
            response.raise_for_status()  # Raise an exception for bad responses

            #check remaining rate limit
            remaining_limit = int(response.headers.get('X-RateLimit-Remaining', 1))
            if remaining_limit < 5:
                reset_time = int(response.headers.get('X-RateLimit-Reset', time.time()))
                sleep_duration = max(0, reset_time - int(time.time()))
                print(f"Rate limit nearing. Sleeping for {sleep_duration} seconds...")
                time.sleep(sleep_duration + 1)
            
            prs = response.json()
            if not prs:
                break
            print(f"Found {len(prs)} PRs on page {page}.")
            all_prs.extend(prs)
            page += 1
            time.sleep(2)  # 2초 대기 후 다음 요청
        except RequestException as e:
            print(f"Error fetching PR list for repository '{repo_name}': {e}")
            break
        
    return all_prs

def get_pr_details(repo_name, pr_number):
    """Fetch the details for a specific PR."""
    pr_detail_url = f"https://api.github.com/repos/AdvancedTechnologyInc/{repo_name}/pulls/{pr_number}"
    try:
        pr_detail_response = requests.get(pr_detail_url, headers=headers)
        pr_detail_response.raise_for_status()  # Raise error for bad responses

        # Check rate limit
        remaining_limit = int(pr_detail_response.headers.get('X-RateLimit-Remaining', 1))
        if remaining_limit < 5:
            reset_time = int(pr_detail_response.headers.get('X-RateLimit-Reset', time.time()))
            sleep_duration = max(0, reset_time - int(time.time()))
            print(f"Rate limit nearing. Sleeping for {sleep_duration} seconds...")
            time.sleep(sleep_duration + 1)

        pr_details = pr_detail_response.json()
        additions = pr_details.get('additions', 0)
        deletions = pr_details.get('deletions', 0)
        total_changes = additions + deletions
        return total_changes
    except RequestException as e:
        print(f"Failed to fetch details for PR #{pr_number}: {e}")
        return 'N/A'

def calculate_merge_time(created_at, closed_at):
    """Calculate the time taken to merge a PR."""
    if created_at and closed_at:
        created_time = datetime.strptime(created_at, "%Y-%m-%dT%H:%M:%SZ")
        closed_time = datetime.strptime(closed_at, "%Y-%m-%dT%H:%M:%SZ")
        return (closed_time - created_time).days
    return 'N/A'

def extract_data_from_prs(prs, repo_name, user_id, start_date=None, end_date=None):
    """
    Extract relevant PR data and include merge/cancel status, filtering by date range.
    
    Args:
        prs (list): List of PRs.
        repo_name (str): Repository name.
        user_id (int): Target user ID.
        start_date (str): Start date in 'YYYY-MM-DD' format (inclusive).
        end_date (str): End date in 'YYYY-MM-DD' format (inclusive).

    Returns:
        list: Filtered and processed PR data.
    """
    data = []

    # Parse the date range
    start_date = datetime.strptime(start_date, "%Y-%m-%d") if start_date else None
    end_date = datetime.strptime(end_date, "%Y-%m-%d") if end_date else None

    for pr in prs:
        # Check if the PR is created by the contributor (user_id)
        if pr['user']['id'] == user_id:
            created_at = datetime.strptime(pr['created_at'], "%Y-%m-%dT%H:%M:%SZ")  # PR creation time
            
            # Filter by date range
            if (start_date and created_at < start_date) or (end_date and created_at > end_date):
                continue  # Skip PRs outside the date range

            pr_title = pr['title']
            pr_number = pr['number']
            pr_link = pr['html_url']
            closed_at_raw = pr['closed_at']
            closed_at = (
                datetime.strptime(closed_at_raw, "%Y-%m-%dT%H:%M:%SZ").strftime("%Y-%m-%d %H:%M:%S")
                if closed_at_raw
                else None
            )
            merged_at = pr['merged_at']

            merge_status = 'Merged' if merged_at else ('Cancelled' if closed_at_raw else 'Open')
            time_to_merge = calculate_merge_time(pr['created_at'], closed_at_raw)
            total_changes = get_pr_details(repo_name, pr_number)

            data.append([
                repo_name, pr_title, pr_number, pr_link,
                created_at.strftime("%Y-%m-%d %H:%M:%S"), closed_at,
                time_to_merge, total_changes, merge_status
            ])

    return data

def save_to_excel(data, output_path):
    """Save PR data to an Excel file with PR Number next to PR Title and hyperlinks on PR Number."""
    print("\nSaving data to Excel...")

    try:
        # Add the 'No.' column by adding index numbers to the data
        numbered_data = [[i + 1] + row[:3] + row[4:] for i, row in enumerate(data)]
        
        # Verify data structure
        if data:
            print(f"Sample row: {data[0]} (Length: {len(data[0])})")

        # Create DataFrame with added 'No.' column
        df = pd.DataFrame(
            numbered_data, 
            columns = ['No.', 'Repository', 'PR Title',
                     'PR No.', 'PR Open Time', 'PR Close Time',
                     'Merge days', 'LOC', 'PR Status']
        )

        # Reorder columns to place PR Number next to PR Title
        column_order = ['No.', 'Repository', 'PR Title',
                     'PR No.', 'PR Open Time', 'PR Close Time',
                     'Merge days', 'LOC', 'PR Status']
        
        if not all(col in df.columns for col in column_order):
            print(f"Invalid column names in column_order: {column_order}. DataFrame columns: {df.columns}")
            return
        df = df[column_order]

        # Write DataFrame to Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='PR List')
            wb = writer.book
            ws = wb['PR List']

            # Add hyperlinks to the 'Number' column
            for row in range(2, len(df) + 2):
                pr_number_cell = ws.cell(row=row, column=4)
                pr_number_cell.value = f"#{data[row-2][2]}"
                pr_number_cell.hyperlink = data[row-2][3]
                pr_number_cell.style = 'Hyperlink'  # Apply hyperlink style
            
            # Auto-adjust column widths # 이거 다시 봐야됨
            MAX_COLUMN_WIDTH = 80
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)  # Get the column letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, MAX_COLUMN_WIDTH)  # Add some padding
                ws.column_dimensions[col_letter].width = adjusted_width

            # Apply alignment to each cell
            for row in ws.iter_rows(min_row=2, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
                for cell in row:
                    if cell.column in (1, 4, 5, 6, 7, 8, 9):
                        cell.alignment = Alignment(horizontal='center')
                    elif cell.column in (2, 3):
                        cell.alignment = Alignment(horizontal='left')
                    else:
                        cell.alignment = Alignment(horizontal='right')

        print(f"PR list has been saved to '{output_path}' with hyperlinks on PR Number.")
    except PermissionError:
        print(f"Permission Error: Unable to write to '{output_path}'. File might be open.")
    except Exception as e:
        print(f"An unexpected error occurred while saving Excel: {e}")

def get_pr_count(repo_name):
    """Check the number of PRs in a repository using Issues API."""
    ensure_rate_limit()
    url = f'https://api.github.com/search/issues?q=repo:AdvancedTechnologyInc/{repo_name}+is:pr'
    try:
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()

        #check remaining rate limit
        remaining_limit = int(response.headers.get('X-RateLimit-Remaining', 1))
        if remaining_limit < 5:
            reset_time = int(response.headers.get('X-RateLimit-Reset', time.time()))
            sleep_duration = max(0, reset_time - int(time.time()))
            print(f"Rate limit nearing. Sleeping for {sleep_duration} seconds...")
            time.sleep(sleep_duration + 1)

        data = response.json()
        return data.get('total_count', 0)  # Total number of PRs
    except RequestException as e:
        print(f"Error checking PR count for repository '{repo_name}': {e}")
        return 0

def main():
    # Excel settings
    # 날짜 범위 설정
    start_date = "2024-06-26"
    end_date = "2024-11-30" # UTC 기준이니 주의!
    repo_excel_path = 'repositorylist_241129.xlsx'
    repo_sheet_name = 'repositories'
    repo_column_letter = 'C'  # Repo name in X column
    contributor_sheet_name = 'contributors'
    contributor_column_letter = 'A'  # Contributor name in A column

    # Step 1: Check rate limit
    get_rate_limit()

    # Step 2: Fetch repository names from Excel sheet
    repos = get_repositories_from_excel(repo_excel_path, repo_sheet_name, repo_column_letter)

    # Step 3: Fetch contributors from Excel sheet
    contributors = get_contributors_from_excel(repo_excel_path, contributor_sheet_name, contributor_column_letter)

    # 사용자 ID 캐시용 딕셔너리
    user_ids = {}

    # Step 4: Fetch PRs and extract data for each contributor
    for contributor in contributors:
        print(f"Processing PRs for contributor '{contributor}' across all repositories...")
        
        # 기여자별 PR 데이터를 누적할 리스트
        contributor_data = []

        # 기여자 ID 캐시에서 확인
        if contributor not in user_ids:
            user_id = get_user_id(contributor)
            if user_id:
                user_ids[contributor] = user_id
            else:
                print(f"Skipping contributor '{contributor}' due to missing user ID.")
                continue  # 사용자 ID를 못 얻으면 다음으로 넘어감
        else:
            user_id = user_ids[contributor]  # 캐시된 사용자 ID 사용

        # 모든 저장소에 대해 PR 데이터 처리
        for repo_name in repos:
            print(f"Checking PR count for repository '{repo_name}'...")
            if repo_name in pr_cache:
                pr_count = get_pr_count(repo_name)
            else:
                pr_count = get_pr_count(repo_name)
                pr_cache[repo_name] = pr_count
                save_cache(pr_cache)

            if pr_count == 0:
                print(f"No PRs found in repository '{repo_name}'. Skipping...\n")
                continue

            # PR이 있는 경우에만 데이터를 가져옴
            prs = get_prs_for_repository(repo_name)
            repo_contributor_data = extract_data_from_prs(prs, repo_name, user_id, start_date, end_date)
            contributor_data.extend(repo_contributor_data)

        # Step 5: 기여자의 모든 데이터를 Excel에 저장
        if contributor_data:
            output_path = f'{contributor}_pr_list.xlsx'
            save_to_excel(contributor_data, output_path)
        else:
            print(f"No PR data found for contributor '{contributor}' in any repository.")

if __name__ == '__main__':
    main()
